"""
Bulk Import API for Campus Africa Student Registration
Handles password-protected Excel files and creates Supabase auth users
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from typing import Optional, List, Dict, Any
from auth import get_current_user, AuthenticatedUser
from supabase_client import supabase
from config import SUPABASE_URL, SUPABASE_SERVICE_KEY
import logging
import uuid
import io
import re
from datetime import datetime
import httpx

# Excel handling
import openpyxl
import msoffcrypto

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/admin/bulk-import", tags=["Bulk Import"])


# ============ Helper Functions ============

def validate_sa_id(id_number: str) -> dict:
    """Validate South African ID number and extract date of birth"""
    if not id_number or len(id_number) != 13:
        return {"valid": False, "error": "ID must be 13 digits"}
    
    if not id_number.isdigit():
        return {"valid": False, "error": "ID must contain only digits"}
    
    # Extract DOB (YYMMDD)
    try:
        yy = int(id_number[0:2])
        mm = int(id_number[2:4])
        dd = int(id_number[4:6])
        
        # Determine century (assume 2000s for years 00-25, 1900s otherwise)
        year = 2000 + yy if yy <= 25 else 1900 + yy
        
        dob = datetime(year, mm, dd)
        dob_str = dob.strftime("%Y-%m-%d")
    except ValueError:
        return {"valid": False, "error": "Invalid date in ID"}
    
    # Extract gender (5000+ = male, <5000 = female)
    gender_digit = int(id_number[6:10])
    gender = "male" if gender_digit >= 5000 else "female"
    
    # Luhn checksum validation
    total = 0
    for i, digit in enumerate(id_number[:-1]):
        d = int(digit)
        if i % 2 == 1:
            d *= 2
            if d > 9:
                d -= 9
        total += d
    
    checksum = (10 - (total % 10)) % 10
    if checksum != int(id_number[-1]):
        return {"valid": False, "error": "Invalid ID checksum"}
    
    return {
        "valid": True,
        "date_of_birth": dob_str,
        "gender": gender
    }


def validate_email(email: str) -> bool:
    """Basic email validation"""
    if not email:
        return False
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, email.strip()))


def normalize_phone(phone: str) -> str:
    """Normalize South African phone number"""
    if not phone:
        return ""
    # Remove spaces and special characters
    phone = re.sub(r'[^\d+]', '', str(phone))
    # Convert to international format if needed
    if phone.startswith('0') and len(phone) == 10:
        phone = '+27' + phone[1:]
    return phone


def parse_date(date_value) -> Optional[str]:
    """Parse date from various formats"""
    if not date_value:
        return None
    
    if isinstance(date_value, datetime):
        return date_value.strftime("%Y-%m-%d")
    
    date_str = str(date_value).strip()
    
    # Try common formats
    formats = [
        "%Y/%m/%d", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y",
        "%Y/%m/%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    
    return None


async def create_supabase_user(email: str, user_data: dict) -> dict:
    """Create a user in Supabase Auth without sending confirmation email"""
    url = f"{SUPABASE_URL}/auth/v1/admin/users"
    
    headers = {
        'apikey': SUPABASE_SERVICE_KEY,
        'Authorization': f'Bearer {SUPABASE_SERVICE_KEY}',
        'Content-Type': 'application/json'
    }
    
    # Create user with auto-confirmed email (no email sent)
    payload = {
        'email': email,
        'email_confirm': True,  # Pre-verify email so no confirmation email is sent
        'user_metadata': {
            'first_name': user_data.get('first_name', ''),
            'last_name': user_data.get('last_name', ''),
            'role': 'patient',
            'id_number': user_data.get('id_number', ''),
            'imported_from': 'campus_africa_bulk',
            'imported_at': datetime.utcnow().isoformat()
        }
    }
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.post(url, json=payload, headers=headers)
            
            if response.status_code in [200, 201]:
                return {"success": True, "user": response.json()}
            elif response.status_code == 422 and "already been registered" in response.text:
                return {"success": False, "error": "Email already registered", "duplicate": True}
            else:
                logger.error(f"Supabase user creation failed: {response.status_code} - {response.text}")
                return {"success": False, "error": f"Auth error: {response.status_code}"}
    except httpx.TimeoutException:
        logger.error(f"Timeout creating user {email}")
        return {"success": False, "error": "Connection timeout to Supabase"}
    except Exception as e:
        logger.error(f"Error creating user {email}: {e}")
        return {"success": False, "error": str(e)}


# ============ API Endpoints ============

@router.post("/preview")
async def preview_import(
    file: UploadFile = File(...),
    password: Optional[str] = Form(None),
    user: AuthenticatedUser = Depends(get_current_user)
):
    """
    Preview the Excel file contents before importing.
    Returns first 10 rows with validation status.
    """
    # Check admin role
    roles = await supabase.select('user_roles', 'role', {'user_id': user.id})
    if not roles or roles[0].get('role') != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Read file
    content = await file.read()
    file_stream = io.BytesIO(content)
    
    # Try to open file (handle password protection)
    try:
        # Try to decrypt if password protected
        if password:
            decrypted = io.BytesIO()
            ms_file = msoffcrypto.OfficeFile(file_stream)
            if ms_file.is_encrypted():
                try:
                    ms_file.load_key(password=password)
                    ms_file.decrypt(decrypted)
                    decrypted.seek(0)
                    workbook = openpyxl.load_workbook(decrypted, read_only=True, data_only=True)
                except Exception:
                    raise HTTPException(status_code=400, detail="Invalid password for encrypted file")
            else:
                file_stream.seek(0)
                workbook = openpyxl.load_workbook(file_stream, read_only=True, data_only=True)
        else:
            # Try without password first
            try:
                workbook = openpyxl.load_workbook(file_stream, read_only=True, data_only=True)
            except Exception:
                # Check if file is encrypted
                file_stream.seek(0)
                ms_file = msoffcrypto.OfficeFile(file_stream)
                if ms_file.is_encrypted():
                    raise HTTPException(
                        status_code=400, 
                        detail="File is password protected. Please provide the password."
                    )
                raise
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error opening Excel file: {e}")
        raise HTTPException(status_code=400, detail=f"Cannot open file: {str(e)}")
    
    # Get first sheet
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="File must have headers and at least one data row")
    
    # Parse headers
    headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    
    # Map expected columns
    column_map = {
        'quadcare account number': 'account_number',
        'account number': 'account_number',
        'title': 'title',
        'first name': 'first_name',
        'firstname': 'first_name',
        'last name': 'last_name',
        'lastname': 'last_name',
        'surname': 'last_name',
        'i.d number': 'id_number',
        'id number': 'id_number',
        'id_number': 'id_number',
        'idnumber': 'id_number',
        'dob': 'date_of_birth',
        'date of birth': 'date_of_birth',
        'gender': 'gender',
        'sex': 'gender',
        'cell': 'phone',
        'phone': 'phone',
        'mobile': 'phone',
        'cellphone': 'phone',
        'email': 'email',
        'e-mail': 'email',
        'employer': 'employer',
        'company': 'employer',
        'occupation': 'occupation',
        'job': 'occupation',
        'status': 'status'
    }
    
    # Map headers to standard names
    mapped_headers = []
    for h in headers:
        mapped = column_map.get(h, h)
        mapped_headers.append(mapped)
    
    # Get existing emails for duplicate check
    existing_users = await supabase.select('profiles', 'email', {})
    existing_emails = {u['email'].lower() for u in existing_users if u.get('email')}
    
    # First pass: Count ALL rows for summary
    total_rows = len(rows) - 1  # Exclude header
    total_new_count = 0
    total_duplicate_count = 0
    total_error_count = 0
    
    for row in rows[1:]:  # All data rows
        row_data = dict(zip(mapped_headers, row))
        email = str(row_data.get('email', '')).strip().lower() if row_data.get('email') else ''
        
        # Only skip if email is missing/invalid OR already in our database
        if not email or not validate_email(email):
            total_error_count += 1
        elif email in existing_emails:
            total_duplicate_count += 1
        else:
            total_new_count += 1
    
    # Second pass: Build preview rows (first 10 data rows)
    preview_rows = []
    
    for row_idx, row in enumerate(rows[1:11], start=2):  # First 10 data rows
        row_data = dict(zip(mapped_headers, row))
        
        # Extract and validate
        email = str(row_data.get('email', '')).strip().lower() if row_data.get('email') else ''
        status = str(row_data.get('status', '')).strip().lower() if row_data.get('status') else ''
        id_number = str(row_data.get('id_number', '')).strip() if row_data.get('id_number') else ''
        
        # Determine import status for this preview row
        validation_errors = []
        import_action = 'import'
        
        # Only skip if email is missing/invalid or already in our database
        if not email:
            import_action = 'error'
            validation_errors.append("Missing email")
        elif not validate_email(email):
            import_action = 'error'
            validation_errors.append("Invalid email format")
        elif email in existing_emails:
            import_action = 'duplicate'
            validation_errors.append("Email already in Quadcare")
        else:
            # Validate ID number if present (warning only, still imports)
            if id_number:
                id_validation = validate_sa_id(id_number)
                if not id_validation['valid']:
                    validation_errors.append(f"ID warning: {id_validation['error']}")
        
        preview_rows.append({
            'row_number': row_idx,
            'account_number': row_data.get('account_number', ''),
            'title': row_data.get('title', ''),
            'first_name': row_data.get('first_name', ''),
            'last_name': row_data.get('last_name', ''),
            'email': email,
            'phone': normalize_phone(str(row_data.get('phone', ''))),
            'id_number': id_number,
            'date_of_birth': parse_date(row_data.get('date_of_birth')),
            'gender': str(row_data.get('gender', '')).lower() if row_data.get('gender') else '',
            'employer': row_data.get('employer', ''),
            'status': status,  # This is Campus Africa status (New/Existing student), not import status
            'import_action': import_action,
            'validation_errors': validation_errors
        })
    
    workbook.close()
    
    return {
        "success": True,
        "file_name": file.filename,
        "total_rows": total_rows,
        "preview_rows": preview_rows,
        "headers_found": mapped_headers,
        "summary": {
            "to_import": total_new_count,
            "duplicates": total_duplicate_count,
            "errors": total_error_count
        }
    }


@router.post("/students")
async def import_students(
    file: UploadFile = File(...),
    password: Optional[str] = Form(None),
    user: AuthenticatedUser = Depends(get_current_user)
):
    """
    Import students from Excel file.
    Creates Supabase auth users and profiles for new students.
    """
    # Check admin role
    roles = await supabase.select('user_roles', 'role', {'user_id': user.id})
    if not roles or roles[0].get('role') != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Read file
    content = await file.read()
    file_stream = io.BytesIO(content)
    
    # Handle password protection
    try:
        if password:
            decrypted = io.BytesIO()
            ms_file = msoffcrypto.OfficeFile(file_stream)
            if ms_file.is_encrypted():
                ms_file.load_key(password=password)
                ms_file.decrypt(decrypted)
                decrypted.seek(0)
                workbook = openpyxl.load_workbook(decrypted, data_only=True)
            else:
                file_stream.seek(0)
                workbook = openpyxl.load_workbook(file_stream, data_only=True)
        else:
            try:
                workbook = openpyxl.load_workbook(file_stream, data_only=True)
            except Exception:
                file_stream.seek(0)
                ms_file = msoffcrypto.OfficeFile(file_stream)
                if ms_file.is_encrypted():
                    raise HTTPException(
                        status_code=400, 
                        detail="File is password protected. Please provide the password."
                    )
                raise
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error opening Excel file: {e}")
        raise HTTPException(status_code=400, detail=f"Cannot open file: {str(e)}")
    
    # Get sheet and parse
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="File must have headers and data")
    
    # Parse headers
    headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    
    column_map = {
        'quadcare account number': 'account_number',
        'account number': 'account_number',
        'title': 'title',
        'first name': 'first_name',
        'firstname': 'first_name',
        'last name': 'last_name',
        'lastname': 'last_name',
        'surname': 'last_name',
        'i.d number': 'id_number',
        'id number': 'id_number',
        'id_number': 'id_number',
        'idnumber': 'id_number',
        'dob': 'date_of_birth',
        'date of birth': 'date_of_birth',
        'gender': 'gender',
        'sex': 'gender',
        'cell': 'phone',
        'phone': 'phone',
        'mobile': 'phone',
        'cellphone': 'phone',
        'email': 'email',
        'e-mail': 'email',
        'employer': 'employer',
        'company': 'employer',
        'occupation': 'occupation',
        'job': 'occupation',
        'status': 'status'
    }
    
    mapped_headers = [column_map.get(h, h) for h in headers]
    
    # Get existing emails
    existing_users = await supabase.select('profiles', 'email', {})
    existing_emails = {u['email'].lower() for u in existing_users if u.get('email')}
    
    # Process all rows
    results = {
        'imported': 0,
        'skipped': 0,
        'errors': 0,
        'duplicates': 0,
        'details': []
    }
    
    try:
        for row_idx, row in enumerate(rows[1:], start=2):
            row_data = dict(zip(mapped_headers, row))
            
            email = str(row_data.get('email', '')).strip().lower() if row_data.get('email') else ''
            # Note: status column is Campus Africa status (New/Existing student), not used for import decisions
            
            # Validate email - only skip if email is invalid
            if not email or not validate_email(email):
                results['errors'] += 1
                results['details'].append({
                    'row': row_idx,
                    'email': email or 'N/A',
                    'status': 'error',
                    'reason': 'Invalid or missing email'
                })
                continue
            
            # Check for duplicates
            if email in existing_emails:
                results['duplicates'] += 1
                results['details'].append({
                    'row': row_idx,
                    'email': email,
                    'status': 'duplicate',
                    'reason': 'Email already exists'
                })
                continue
            
            # Prepare user data
            first_name = str(row_data.get('first_name', '')).strip() if row_data.get('first_name') else ''
            last_name = str(row_data.get('last_name', '')).strip() if row_data.get('last_name') else ''
            id_number = str(row_data.get('id_number', '')).strip() if row_data.get('id_number') else ''
            phone = normalize_phone(str(row_data.get('phone', ''))) if row_data.get('phone') else ''
            
            # Parse DOB from ID or directly
            dob = None
            gender = str(row_data.get('gender', '')).lower() if row_data.get('gender') else None
            
            if id_number:
                id_validation = validate_sa_id(id_number)
                if id_validation['valid']:
                    dob = id_validation['date_of_birth']
                    gender = id_validation['gender']
            
            if not dob:
                dob = parse_date(row_data.get('date_of_birth'))
            
            user_data = {
                'first_name': first_name,
                'last_name': last_name,
                'id_number': id_number,
                'phone': phone,
                'date_of_birth': dob,
                'gender': gender,
                'employer': row_data.get('employer', 'Campus Africa'),
                'account_number': row_data.get('account_number', '')
            }
            
            # Create Supabase auth user
            auth_result = await create_supabase_user(email, user_data)
            
            if not auth_result['success']:
                if auth_result.get('duplicate'):
                    results['duplicates'] += 1
                    existing_emails.add(email)  # Add to local set
                else:
                    results['errors'] += 1
                results['details'].append({
                    'row': row_idx,
                    'email': email,
                    'status': 'error' if not auth_result.get('duplicate') else 'duplicate',
                    'reason': auth_result['error']
                })
                continue
            
            # Get user ID from auth response
            new_user_id = auth_result['user']['id']
            
            # Create/update profile
            profile_data = {
                'id': new_user_id,
                'email': email,
                'first_name': first_name,
                'last_name': last_name,
                'phone': phone,
                'id_number': id_number,
                'date_of_birth': dob,
                'created_at': datetime.utcnow().isoformat(),
                'updated_at': datetime.utcnow().isoformat()
            }
            
            # Insert profile
            profile_result = await supabase.insert('profiles', profile_data)
            
            if not profile_result:
                logger.error(f"Failed to create profile for {email}")
                results['errors'] += 1
                results['details'].append({
                    'row': row_idx,
                    'email': email,
                    'status': 'error',
                    'reason': 'Failed to create profile (auth user created)'
                })
                continue
            
            # Create user role
            role_data = {
                'id': str(uuid.uuid4()),
                'user_id': new_user_id,
                'role': 'patient'
            }
            
            await supabase.insert('user_roles', role_data)
            
            # Success
            results['imported'] += 1
            existing_emails.add(email)  # Prevent duplicates in same batch
            results['details'].append({
                'row': row_idx,
                'email': email,
                'name': f"{first_name} {last_name}",
                'status': 'imported',
                'reason': 'Successfully created'
            })
    
    except Exception as e:
        logger.error(f"Bulk import error at row processing: {e}")
        # Return partial results if we got interrupted
        results['details'].append({
            'row': 'N/A',
            'email': 'N/A',
            'status': 'error',
            'reason': f'Import interrupted: {str(e)}'
        })
    
    workbook.close()
    
    logger.info(f"Bulk import completed: {results['imported']} imported, {results['skipped']} skipped, {results['duplicates']} duplicates, {results['errors']} errors")
    
    return {
        "success": True,
        "summary": {
            "total_processed": len(rows) - 1,
            "imported": results['imported'],
            "skipped": results['skipped'],
            "duplicates": results['duplicates'],
            "errors": results['errors']
        },
        "details": results['details'][:100]  # Limit details to first 100 for response size
    }


@router.get("/template")
async def get_import_template(
    user: AuthenticatedUser = Depends(get_current_user)
):
    """
    Get information about the expected Excel template format.
    """
    # Check admin role
    roles = await supabase.select('user_roles', 'role', {'user_id': user.id})
    if not roles or roles[0].get('role') != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    return {
        "expected_columns": [
            {"name": "Quadcare Account Number", "required": False, "description": "Student account ID (e.g., BM-0001)"},
            {"name": "Title", "required": False, "description": "Mr, Ms, Mrs, etc."},
            {"name": "First Name", "required": True, "description": "Student's first name"},
            {"name": "Last Name", "required": True, "description": "Student's last name"},
            {"name": "I.D Number", "required": False, "description": "SA ID (13 digits) or Passport"},
            {"name": "DOB", "required": False, "description": "Date of birth (YYYY/MM/DD)"},
            {"name": "Gender", "required": False, "description": "male/female"},
            {"name": "Cell", "required": False, "description": "Phone number"},
            {"name": "Email", "required": True, "description": "Email address (must be unique)"},
            {"name": "Employer", "required": False, "description": "Company/Institution name"},
            {"name": "Occupation", "required": False, "description": "Student, Staff, etc."},
            {"name": "Status", "required": False, "description": "ExistingUser = skip import"}
        ],
        "notes": [
            "Rows with Status='ExistingUser' will be skipped",
            "Duplicate emails will be skipped",
            "SA ID numbers will be validated and DOB/gender extracted",
            "Password-protected Excel files are supported (provide password)",
            "Students can use 'Forgot Password' to set their login password"
        ]
    }
