"""
Script to update profiles of bulk-imported users with their COMPLETE data from the Excel file.
Run this to fix profiles that were imported but missing data.

Usage:
    python update_imported_profiles.py --file /path/to/excel.xlsx --password yourpassword
    python update_imported_profiles.py --file /path/to/excel.xlsx  # If not password protected
"""
import asyncio
import httpx
import sys
import os
import io
import re
from datetime import datetime

# Excel handling
import openpyxl
import msoffcrypto

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import SUPABASE_URL, SUPABASE_SERVICE_KEY


def normalize_phone(phone: str) -> str:
    """Normalize South African phone number"""
    if not phone:
        return None
    phone = re.sub(r'[^\d+]', '', str(phone))
    if phone.startswith('0') and len(phone) == 10:
        phone = '+27' + phone[1:]
    return phone if phone else None


def parse_date(date_value) -> str:
    """Parse date from various formats"""
    if not date_value:
        return None
    
    if isinstance(date_value, datetime):
        return date_value.strftime("%Y-%m-%d")
    
    date_str = str(date_value).strip()
    
    formats = [
        "%Y/%m/%d", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y",
        "%Y/%m/%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    
    return None


def validate_sa_id(id_number: str) -> dict:
    """Validate South African ID and extract DOB/gender"""
    if not id_number or len(id_number) != 13 or not id_number.isdigit():
        return {"valid": False}
    
    try:
        yy = int(id_number[0:2])
        mm = int(id_number[2:4])
        dd = int(id_number[4:6])
        year = 2000 + yy if yy <= 25 else 1900 + yy
        dob = datetime(year, mm, dd)
        gender_digit = int(id_number[6:10])
        gender = "male" if gender_digit >= 5000 else "female"
        return {"valid": True, "date_of_birth": dob.strftime("%Y-%m-%d"), "gender": gender}
    except:
        return {"valid": False}


async def get_auth_users_by_email():
    """Get all auth users mapped by email"""
    users_by_email = {}
    page = 1
    
    async with httpx.AsyncClient(timeout=30.0) as client:
        while True:
            response = await client.get(
                f"{SUPABASE_URL}/auth/v1/admin/users",
                params={'page': page, 'per_page': 100},
                headers={
                    'apikey': SUPABASE_SERVICE_KEY,
                    'Authorization': f'Bearer {SUPABASE_SERVICE_KEY}'
                }
            )
            
            if response.status_code != 200:
                break
            
            data = response.json()
            batch = data.get('users', [])
            
            if not batch:
                break
            
            for user in batch:
                email = user.get('email', '').lower()
                if email:
                    users_by_email[email] = user
            
            print(f"Fetched {len(users_by_email)} auth users...")
            
            if len(batch) < 100:
                break
            page += 1
    
    return users_by_email


async def update_profile(user_id: str, data: dict):
    """Update a profile in Supabase"""
    url = f"{SUPABASE_URL}/rest/v1/profiles?id=eq.{user_id}"
    
    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.patch(
            url,
            json=data,
            headers={
                'apikey': SUPABASE_SERVICE_KEY,
                'Authorization': f'Bearer {SUPABASE_SERVICE_KEY}',
                'Content-Type': 'application/json',
                'Prefer': 'return=representation'
            }
        )
        return response.status_code == 200


async def main():
    import argparse
    parser = argparse.ArgumentParser(description='Update profiles for bulk-imported users')
    parser.add_argument('--file', required=True, help='Path to Excel file')
    parser.add_argument('--password', default=None, help='Excel file password if protected')
    args = parser.parse_args()
    
    print("=" * 60)
    print("UPDATE IMPORTED PROFILES - FULL DATA")
    print("=" * 60)
    
    # Open Excel file
    print(f"\nOpening Excel file: {args.file}")
    with open(args.file, 'rb') as f:
        file_stream = io.BytesIO(f.read())
    
    if args.password:
        decrypted = io.BytesIO()
        ms_file = msoffcrypto.OfficeFile(file_stream)
        if ms_file.is_encrypted():
            ms_file.load_key(password=args.password)
            ms_file.decrypt(decrypted)
            decrypted.seek(0)
            workbook = openpyxl.load_workbook(decrypted, data_only=True)
        else:
            file_stream.seek(0)
            workbook = openpyxl.load_workbook(file_stream, data_only=True)
    else:
        workbook = openpyxl.load_workbook(file_stream, data_only=True)
    
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    
    # Parse headers - be flexible with column names
    headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    
    column_map = {
        'quadcare account number': 'account_number',
        'account number': 'account_number',
        'title': 'title',
        'first name': 'first_name',
        'firstname': 'first_name',
        'last name': 'last_name',
        'lastname': 'last_name',
        'surname': 'last_name',
        'i.d number': 'id_number',
        'id number': 'id_number',
        'id_number': 'id_number',
        'idnumber': 'id_number',
        'dob': 'date_of_birth',
        'date of birth': 'date_of_birth',
        'gender': 'gender',
        'sex': 'gender',
        'cell': 'phone',
        'phone': 'phone',
        'mobile': 'phone',
        'cellphone': 'phone',
        'email': 'email',
        'e-mail': 'email',
        'employer': 'employer',
        'company': 'employer',
        'occupation': 'occupation',
        'job': 'occupation',
        'status': 'import_status'
    }
    
    mapped_headers = [column_map.get(h, h) for h in headers]
    
    print(f"Found columns: {[h for h in mapped_headers if not h.startswith('col_')]}")
    print(f"Found {len(rows) - 1} data rows")
    
    # Get auth users
    print("\nFetching auth users from Supabase...")
    auth_users = await get_auth_users_by_email()
    print(f"Found {len(auth_users)} auth users")
    
    # Process rows
    updated = 0
    skipped = 0
    not_found = 0
    
    for row_idx, row in enumerate(rows[1:], start=2):
        row_data = dict(zip(mapped_headers, row))
        
        email = str(row_data.get('email', '')).strip().lower() if row_data.get('email') else ''
        
        if not email:
            skipped += 1
            continue
        
        # Check if user exists in auth
        auth_user = auth_users.get(email)
        if not auth_user:
            not_found += 1
            continue
        
        user_id = auth_user['id']
        
        # Extract ALL fields
        first_name = str(row_data.get('first_name', '')).strip() if row_data.get('first_name') else None
        last_name = str(row_data.get('last_name', '')).strip() if row_data.get('last_name') else None
        id_number = str(row_data.get('id_number', '')).strip() if row_data.get('id_number') else None
        phone = normalize_phone(str(row_data.get('phone', ''))) if row_data.get('phone') else None
        title = str(row_data.get('title', '')).strip() if row_data.get('title') else None
        account_number = str(row_data.get('account_number', '')).strip() if row_data.get('account_number') else None
        employer = str(row_data.get('employer', '')).strip() if row_data.get('employer') else 'Campus Africa'
        occupation = str(row_data.get('occupation', '')).strip() if row_data.get('occupation') else None
        import_status = str(row_data.get('import_status', '')).strip() if row_data.get('import_status') else None
        
        # Parse DOB and gender from ID if available
        dob = None
        gender = str(row_data.get('gender', '')).lower() if row_data.get('gender') else None
        
        if id_number:
            # Clean ID number (remove spaces, etc)
            id_number = re.sub(r'[^\d]', '', id_number)
            id_validation = validate_sa_id(id_number)
            if id_validation.get('valid'):
                dob = id_validation['date_of_birth']
                gender = id_validation['gender']
        
        if not dob:
            dob = parse_date(row_data.get('date_of_birth'))
        
        # Build profile data - only include non-None values
        profile_data = {
            'updated_at': datetime.utcnow().isoformat()
        }
        
        if first_name: profile_data['first_name'] = first_name
        if last_name: profile_data['last_name'] = last_name
        if phone: profile_data['phone'] = phone
        if id_number: profile_data['id_number'] = id_number
        if dob: profile_data['date_of_birth'] = dob
        if gender: profile_data['gender'] = gender
        if title: profile_data['title'] = title
        if account_number: profile_data['account_number'] = account_number
        if employer: profile_data['employer'] = employer
        if occupation: profile_data['occupation'] = occupation
        if import_status: profile_data['import_status'] = import_status
        
        # Update profile
        success = await update_profile(user_id, profile_data)
        
        if success:
            updated += 1
            if updated % 100 == 0:
                print(f"  Updated {updated} profiles...")
        else:
            skipped += 1
    
    workbook.close()
    
    print("\n" + "=" * 60)
    print("UPDATE COMPLETE")
    print(f"  - Updated: {updated}")
    print(f"  - Skipped: {skipped}")
    print(f"  - Not found in auth: {not_found}")
    print("=" * 60)


if __name__ == '__main__':
    asyncio.run(main())
